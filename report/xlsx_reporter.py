"""Excel reporter for validation findings."""

from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Union

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from engines.base import Engine, Finding, Severity


class XLSXReporter:
    """Generate Excel output for validation findings."""

    # Severity fill colors
    SEVERITY_FILLS = {
        Severity.FOUT: PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"),
        Severity.WAARSCHUWING: PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
        Severity.INFO: PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid"),
    }

    # Column definitions
    COLUMNS = [
        ("severity", "Severity", 15),
        ("engine", "Engine", 10),
        ("code", "Code", 10),
        ("regeltype", "Regeltype", 25),
        ("contract", "Contract", 15),
        ("branche", "Branche", 10),
        ("entiteit", "Entiteit", 10),
        ("label", "Label", 20),
        ("waarde", "Waarde", 30),
        ("omschrijving", "Omschrijving", 50),
        ("verwacht", "Verwacht", 40),
        ("bron", "Bron", 25),
    ]

    def __init__(self):
        self.wb: Optional[Workbook] = None

    def write(
        self,
        findings: List[Finding],
        output_path: Union[str, Path],
        source_file: str = "",
    ) -> None:
        """Write Excel report to file."""
        self.wb = Workbook()

        # Remove default sheet
        default_sheet = self.wb.active
        if default_sheet:
            self.wb.remove(default_sheet)

        # Create sheets
        self._create_summary_sheet(findings, source_file)
        self._create_all_findings_sheet(findings)
        self._create_by_severity_sheets(findings)
        self._create_by_contract_sheet(findings)

        # Save workbook
        path = Path(output_path)
        path.parent.mkdir(parents=True, exist_ok=True)
        self.wb.save(str(path))

    def _create_summary_sheet(self, findings: List[Finding], source_file: str) -> None:
        """Create summary sheet."""
        ws = self.wb.create_sheet("Samenvatting")

        # Title
        ws["A1"] = "SIVI AFD XML Validator - Rapport"
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells("A1:D1")

        # Metadata
        ws["A3"] = "Bronbestand:"
        ws["B3"] = source_file or "(onbekend)"
        ws["A4"] = "Datum/tijd:"
        ws["B4"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws["A5"] = "Totaal bevindingen:"
        ws["B5"] = len(findings)

        # Summary by severity
        ws["A7"] = "Per Severity"
        ws["A7"].font = Font(bold=True)

        row = 8
        for severity in [Severity.FOUT, Severity.WAARSCHUWING, Severity.INFO]:
            count = sum(1 for f in findings if f.severity == severity)
            ws[f"A{row}"] = severity.value
            ws[f"B{row}"] = count
            ws[f"A{row}"].fill = self.SEVERITY_FILLS[severity]
            row += 1

        # Summary by engine
        ws[f"A{row + 1}"] = "Per Engine"
        ws[f"A{row + 1}"].font = Font(bold=True)

        row += 2
        engine_names = {
            Engine.SCHEMA: "Engine 1: Schema",
            Engine.RULES: "Engine 2: Business Rules",
            Engine.LLM: "Engine 3: LLM Semantiek",
        }
        for engine, name in engine_names.items():
            count = sum(1 for f in findings if f.engine == engine)
            ws[f"A{row}"] = name
            ws[f"B{row}"] = count
            row += 1

        # Auto-fit column widths
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 40

    def _create_all_findings_sheet(self, findings: List[Finding]) -> None:
        """Create sheet with all findings."""
        ws = self.wb.create_sheet("Alle bevindingen")
        self._write_findings_table(ws, findings)

    def _create_by_severity_sheets(self, findings: List[Finding]) -> None:
        """Create sheets per severity level."""
        for severity in [Severity.FOUT, Severity.WAARSCHUWING, Severity.INFO]:
            severity_findings = [f for f in findings if f.severity == severity]
            if severity_findings:
                ws = self.wb.create_sheet(severity.value)
                self._write_findings_table(ws, severity_findings)

    def _create_by_contract_sheet(self, findings: List[Finding]) -> None:
        """Create sheet grouped by contract."""
        ws = self.wb.create_sheet("Per contract")

        # Group by contract
        by_contract: Dict[str, List[Finding]] = defaultdict(list)
        for f in findings:
            by_contract[f.contract].append(f)

        row = 1
        for contract in sorted(by_contract.keys()):
            contract_findings = by_contract[contract]

            # Contract header
            ws[f"A{row}"] = f"Contract: {contract}"
            ws[f"A{row}"].font = Font(bold=True, size=12)
            ws.merge_cells(f"A{row}:L{row}")
            row += 1

            # Write findings table
            self._write_findings_table(ws, contract_findings, start_row=row)
            row += len(contract_findings) + 2  # +2 for header and spacing

    def _write_findings_table(
        self, ws, findings: List[Finding], start_row: int = 1
    ) -> None:
        """Write findings as a table."""
        if not findings:
            return

        # Header row
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for col_idx, (field, header, width) in enumerate(self.COLUMNS, 1):
            cell = ws.cell(row=start_row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Data rows
        for row_idx, finding in enumerate(findings, start_row + 1):
            for col_idx, (field, _, _) in enumerate(self.COLUMNS, 1):
                value = getattr(finding, field, "")
                if hasattr(value, "value"):
                    value = value.value

                cell = ws.cell(row=row_idx, column=col_idx, value=str(value))
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top", wrap_text=True)

                # Apply severity color to first column
                if col_idx == 1:
                    cell.fill = self.SEVERITY_FILLS.get(finding.severity, PatternFill())


def report_to_xlsx(
    findings: List[Finding],
    output_path: Union[str, Path],
    source_file: str = "",
) -> None:
    """Convenience function to write Excel report."""
    reporter = XLSXReporter()
    reporter.write(findings, output_path, source_file)
